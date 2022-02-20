from openpyxl import load_workbook
from openpyxl.styles import Font
from selenium.webdriver.chrome.service import Service
import xlsxwriter
import datetime

CHROME_DRIVER_PATH = "C:\DRIVERS\chromedriver_win32\chromedriver.exe"
service = Service(CHROME_DRIVER_PATH)


class FillForms:
    def __init__(self, locations, prices, links):
        self.locations = locations
        self.prices = prices
        self.links = links
        self.blue_font = Font(color='000000FF')     # Set blue fonts (for url links)
        # Get the current date and time
        self.time_stamp = datetime.datetime.now().strftime("%d/""%B/""%Y ""%X")  # Format the date and time as dd/Month/yyyy hh:mm:ss (eg 19/February/2022 16:05:33)

    def fill_excel(self):

        try:        # Open the existing workbook and worksheet if they exist
            self.workbook = load_workbook('Rightmove Houses.xlsx')
            self.worksheet = self.workbook['Properties']
        except FileNotFoundError:        # If they don't exist create them and open
            self.workbook = xlsxwriter.Workbook('Rightmove Houses.xlsx')
            self.worksheet = self.workbook.add_worksheet("Properties")

            # Widen columns
            self.worksheet.set_column('A:A', 25)
            self.worksheet.set_column('B:B', 50)
            self.worksheet.set_column('C:C', 10)
            self.worksheet.set_column('D:D', 80)

            # Set columns' names in bold
            format_bold = self.workbook.add_format({'bold': True})
            self.worksheet.write('A1', 'Timestamp', format_bold)
            self.worksheet.write('B1', 'Locations', format_bold)
            self.worksheet.write('C1', 'Prices', format_bold)
            self.worksheet.write('D1', 'Links', format_bold)

            self.workbook.close()

            self.workbook = load_workbook('Rightmove Houses.xlsx')
            self.worksheet = self.workbook['Properties']

        # Identify the number of occupied rows
        last_row = self.worksheet.max_row
        if last_row == 1:   # Make the 2nd row empty if not already done so
            last_row += 1

        # Fill in  the timestamps
        timestamp_row = last_row + 1
        for _ in range(len(self.locations)):
            self.worksheet[f'A{timestamp_row}'].value = self.time_stamp
            timestamp_row += 1

        # Fill locations in
        location_row = last_row + 1
        for location in self.locations:
            self.worksheet[f'B{location_row}'].value = location
            location_row += 1

        # # Fill prices in
        price_row = last_row + 1
        for price in self.prices:
            self.worksheet[f'C{price_row}'].value = price
            price_row += 1

        # # Fill links in
        link_row = last_row + 1
        for link in self.links:
            self.worksheet[f'D{link_row}'].hyperlink = link
            self.worksheet[f'D{link_row}'].font = self.blue_font
            link_row += 1

        self.workbook.save('Rightmove Houses.xlsx')
